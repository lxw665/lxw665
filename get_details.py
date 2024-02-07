from DrissionPage import ChromiumPage
import time
from openpyxl import load_workbook
import pandas as pd
import os


def drop_down(page):
    """执行页面向下滚动的操作，实际上就等同于在浏览器的控制台上执行一串js代码"""
    for x in range(1, 8, 2):
        # 延时操作
        time.sleep(1)
        # 1/5 3/5 5/5
        j = x / 9
        # document.documentElement.scrollTop 制定滚动条位置
        # document.documentElement.scrollHeight 获取浏览器页面的最大高度
        js = 'document.documentElement.scrollTop = document.documentElement.scrollHeight * %f' % j
        # 运行 js 代码
        page.run_js(js)


def save_to_xls(excel_file, rowData):
    # 如果文件不存在，创建一个空的Excel文件
    if not os.path.exists(excel_file):
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 创建一个空的DataFrame作为占位符，以便创建Excel文件
            empty_df = pd.DataFrame()
            empty_df.to_excel(writer, sheet_name='Sheet1', index=False)
    # 使用openpyxl打开现有的Excel文件
    wb = load_workbook(excel_file)
    ws = wb['Sheet1']
    # 在最后一行下方追加新行
    ws.append(rowData)
    # 保存Excel文件
    wb.save(excel_file)


save_path = "detail_info.xlsx"
t_title = ["pic_url_str", "price", "bedrooms_num", "Bathrooms", "receptions", "address", "position", "features_section",
           "floor_plan", "tenure"]
save_to_xls(save_path, t_title)
# 创建页面对象，并启动或接管浏览器
page = ChromiumPage()
# 指定Excel文件的路径
excel_file_path = 'base_info.xlsx'
# 加载工作簿
workbook = load_workbook(excel_file_path)
# 选择工作表，这里假设你想要读取的第一个工作表
sheet = workbook.active  # 或者使用 workbook.get_sheet_by_name('Sheet1') 如果你知道工作表的名称
# 遍历工作表的所有行
for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):  # 从第二行开始，因为第一行通常是表头
    # 输出每一行的第一个值
    detail_link = row[0]
    print(detail_link)
    page.get(detail_link)
    drop_down(page)
    position = ""
    try:
        position = page.ele("@data-testid=static-map-container").ele("tag:source").attr("srcset").split("center=", 1)[1].split(
            "&maptype=", 1)[0]
    except:
        pass
    floor_plan, features_section = "", ""
    try:
        floor_plan = page.ele("@data-name=floorplan-item").ele('tag:source').attr("srcset").split(":p", 1)[0]
    except:
        pass
    try:
        features_section = page.ele("@data-testid=page_features_section").text
    except:
        pass
    tenure = ""
    try:
        desc_list = page.eles('xpath://*[@id="main-content"]/div/div[2]/div[2]/div[2]/div[5]/div')
        # tenure
        for item in desc_list:
            if "Tenure" in item.text:
                tenure = item.text.split(":", 1)[1].strip()
    except:
        pass

    data_list = [row[1], row[2], row[3], row[4], row[5], row[6], position, features_section, floor_plan, tenure]
    save_to_xls(save_path, data_list)
